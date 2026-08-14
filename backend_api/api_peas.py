# api_peas.py
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_, or_, distinct
from calendar import monthrange
from fastapi import File, UploadFile, Form
from botocore.config import Config
import pandas as pd
from io import BytesIO
from pydantic import BaseModel
from fastapi.responses import FileResponse
import openpyxl
import calendar
from fastapi.responses import StreamingResponse

# Importa tus módulos locales (ajusta los puntos si tu estructura es distinta)
from . import models, schemas
from .database import get_db
from datetime import datetime, date, timedelta
import pytz
import os
import uuid
import boto3
from . import security
from . import models
from . import schemas

router = APIRouter(
    prefix="/api/peas",
    tags=["PEAS Asistencia"]
)

s3_client = boto3.client(
   's3',
    endpoint_url=os.getenv('B2_ENDPOINT'),
    aws_access_key_id=os.getenv('B2_KEY_ID'),
    aws_secret_access_key=os.getenv('B2_APPLICATION_KEY'),
    config=Config(
        signature_version='s3v4',
        s3={'addressing_style': 'virtual'}
    )
)

class RechazoRequest(BaseModel):
    observaciones: str

import pytz
mx_tz = pytz.timezone('America/Mexico_City')
utc_tz = pytz.utc

def formatear_fecha_local(fecha_utc):
    if not fecha_utc:
        return "Sin fecha"
    if fecha_utc.tzinfo is None:
        fecha_utc = utc_tz.localize(fecha_utc)
    return fecha_utc.astimezone(mx_tz).strftime("%Y-%m-%d %H:%M")

@router.post("/usuarios-acceso", response_model=schemas.UsuarioAccesoResponse, tags=["Gestión de Accesos"])
async def registrar_usuario_acceso(
    usuario: schemas.UsuarioAccesoCreate, 
    db: Session = Depends(get_db),
    # Opcional: Descomenta la siguiente línea si quieres que SOLO los admins puedan crear estos accesos
    # current_user: models.User = Depends(security.get_current_active_user) 
):
    # 1. Verificar si el ID IMSS realmente existe en tu padrón de doctores
    doctor_existe = db.query(models.Doctor).filter(models.Doctor.id_imss == usuario.id_imss).first()
    if not doctor_existe:
        raise HTTPException(status_code=404, detail="El ID IMSS no existe en el registro principal de médicos.")

    # 2. Verificar si este médico ya tiene una cuenta de acceso creada
    cuenta_existe = db.query(models.UsuarioAcceso).filter(models.UsuarioAcceso.id_imss == usuario.id_imss).first()
    if cuenta_existe:
        raise HTTPException(status_code=400, detail="Este usuario ya tiene una cuenta de acceso.")

    # 3. Hashear la contraseña y guardar en la nueva tabla
    nuevo_acceso = models.UsuarioAcceso(
        id_imss=usuario.id_imss,
        hashed_password=security.get_password_hash(usuario.password), # ¡Encriptación vital!
        rol=usuario.rol,
        estatus=usuario.estatus
    )
    
    db.add(nuevo_acceso)
    db.commit()
    db.refresh(nuevo_acceso)
    
    return nuevo_acceso

@router.post("/reporte-quincenal/previsualizar-excel", tags=["Encargado Unidad"])
async def previsualizar_excel_asistencia(
    anio: int = Form(...),
    mes: int = Form(...),
    quincena: int = Form(...),
    archivo: UploadFile = File(...)):

    if not archivo.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="El archivo debe ser un formato de Excel (.xlsx)")

    fecha_ini = date(anio, mes, 1) if quincena == 1 else date(anio, mes, 16)
    fecha_fin = date(anio, mes, 15) if quincena == 1 else date(anio, mes, monthrange(anio, mes)[1])

    try:
        contents = await archivo.read()
        df = pd.read_excel(BytesIO(contents))
        df = df.dropna(how='all')

        asistencias_extraidas = []
        dias_validos = 0

        for index, row in df.iterrows():    
            if pd.isna(row.get('FECHA (DD/MM/AAAA)')):
                continue

            fecha_cruda = pd.to_datetime(row['FECHA (DD/MM/AAAA)'], dayfirst=True)

            if not (fecha_ini <= fecha_cruda.date() <= fecha_fin):
                raise HTTPException(
                    status_code=400, 
                    detail=f"Error en fila {index + 2}: La fecha {fecha_cruda.strftime('%d/%m/%Y')} no pertenece a la Quincena {quincena} del mes {mes}/{anio}. Por favor, corrige el Excel."
                )
            
            fecha_str = fecha_cruda.strftime('%Y-%m-%d')
            turno = str(row.get('TIPO_TURNO', 'No especificado')).strip()
            
            entrada_val = row.get('HORA_ENTRADA (HH:MM)')
            salida_val = row.get('HORA_SALIDA (HH:MM)')

            # Si el médico puso horas, las convertimos a string. Si lo dejó vacío, ponemos "--:--"
            hora_entrada = str(entrada_val).strip() if pd.notnull(entrada_val) and str(entrada_val).strip() != "" else "--:--"
            hora_salida = str(salida_val).strip() if pd.notnull(salida_val) and str(salida_val).strip() != "" else "--:--"

            if hora_entrada != "--:--" or hora_salida != "--:--":
                dias_validos += 1

            asistencias_extraidas.append({
                "fecha": fecha_str,
                "turno": turno,
                "entrada": hora_entrada,
                "salida": hora_salida,
                "observaciones": str(row.get('OBSERVACIONES', '')).strip() if pd.notnull(row.get('OBSERVACIONES')) else ""
            })

        return {
            "mensaje": "Excel procesado exitosamente",
            "total_dias_registrados": dias_validos,
            "detalle_asistencias": asistencias_extraidas
        }

    except HTTPException as http_exc:
        raise http_exc 
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno al leer el Excel. Verifica que tenga las columnas correctas. Error: {str(e)}")

# 1. Endpoint para traer los reportes pendientes de un Estado
@router.get("/coordinador/reportes-pendientes/{entidad}", tags=["Coordinador Estatal"])
async def obtener_reportes_pendientes_estado(entidad: str, db: Session = Depends(get_db)):
    entidad_limpia = entidad.upper().strip()

    mapa_entidades = {
        "BAJA CALIFORNIA": "BC",
        "BAJA CALIFORNIA SUR": "BCS",
        "CAMPECHE": "CAMP",
        "CIUDAD DE MEXICO": "CDMX",
        "CHIAPAS": "CHIS",
        "COLIMA": "COL",
        "GUERRERO": "GRO",
        "HIDALGO": "HGO",
        "MEXICO": "MEX",
        "MICHOACAN": "MICH",
        "MORELOS": "MOR",
        "NAYARIT": "NAY",
        "OAXACA": "OAX",
        "PUEBLA": "PUE",
        "QUINTANA ROO": "QROO",
        "SINALOA": "SIN",
        "SAN LUIS POTOSI": "SLP",
        "SONORA": "SON",
        "TABASCO": "TAB",
        "TAMAULIPAS": "TAMPS",
        "TLAXCALA": "TLAX",
        "VERACRUZ": "VER",
        "YUCATAN": "YUC",
        "ZACATECAS": "ZAC"
    }
    entidad_query = mapa_entidades.get(entidad_limpia, entidad_limpia)

    # REGLA DE ORO: Solo traemos reportes cuyo estado sea explícitamente PENDIENTE.
    # Si fue RECHAZADO, ya no entrará aquí, aunque no esté en la tabla de validados.
    reportes = db.query(models.ReporteQuincenal, models.Doctor)\
        .join(models.Doctor, models.ReporteQuincenal.id_imss == models.Doctor.id_imss)\
        .filter(models.Doctor.entidad == entidad_query)\
        .filter(models.ReporteQuincenal.estado == models.EstadoReporte.PENDIENTE)\
        .all()
        
    resultado = []
    for reporte, doctor in reportes:
        resultado.append({
            "id_reporte": reporte.id,
            "id_imss": doctor.id_imss,
            "medico": f"{doctor.nombre} {doctor.apellido_paterno or ''} {doctor.apellido_materno or ''}".strip(),
            "especialidad": doctor.especialidad or "No especificada",
            "turno": doctor.turno or "Matutino",
            "clues": doctor.clues or "Sin CLUES",
            "unidad": doctor.nombre_unidad or "Sin Unidad",
            "quincena": reporte.quincena,
            "url_pdf": reporte.url_documento,
            "subido_por": reporte.subido_por,
            "dias": reporte.total_dias
        })
        
    return resultado

# 2. Endpoint para Validar y Guardar en la nueva tabla
@router.post("/coordinador/validar-reporte", tags=["Coordinador Estatal"])
async def validar_reporte_estatal(
    datos_validacion: dict, # Aquí recibiremos los datos desde React
    db: Session = Depends(get_db)
):
    # Verificamos que no esté validado ya
    existe = db.query(models.BitacoraEstatalValidada).filter(
        models.BitacoraEstatalValidada.id_reporte_quincenal == datos_validacion["id_reporte"]
    ).first()
    
    if existe:
        raise HTTPException(status_code=400, detail="Este documento ya fue validado.")

    nuevo_registro = models.BitacoraEstatalValidada(
        id_reporte_quincenal=datos_validacion["id_reporte"],
        id_imss=datos_validacion["id_imss"],
        quincena_validada=datos_validacion["quincena"],
        profesional_salud=datos_validacion["medico"],
        especialidad=datos_validacion["especialidad"],
        turno=datos_validacion["turno"],
        clues_ib=datos_validacion["clues"],
        unidad_medica=datos_validacion["unidad"],
        dias_participacion=datos_validacion["dias_participacion"], # Lo pasará el coordinador
        entidad=datos_validacion["entidad"],
        validado_por=datos_validacion["validado_por"]
    )
    
    db.add(nuevo_registro)
    reporte_original = db.query(models.ReporteQuincenal).filter(models.ReporteQuincenal.id == datos_validacion["id_reporte"]).first()
    if reporte_original:
        reporte_original.estado = "aprobado"
    db.commit()
    
    return {"mensaje": "Documento validado e insertado en la Bitácora Estatal"}

@router.post("/reporte-quincenal/subir", tags=["Encargado Unidad"])
async def subir_reporte_y_excel(
    id_imss: str = Form(...),
    anio: int = Form(...),
    mes: int = Form(...),
    quincena: int = Form(...),
    subido_por: str = Form(...),
    archivo_pdf: UploadFile = File(...),
    archivo_excel: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    # 1. Validaciones básicas
    if not archivo_pdf.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="El primer archivo debe ser un PDF firmado.")
    if not archivo_excel.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="El segundo archivo debe ser un Excel (.xlsx).")

    doctor = db.query(models.Doctor).filter(models.Doctor.id_imss == id_imss).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Médico no encontrado")

    # 2. Subir ambos archivos a Backblaze B2
    periodo_str = f"{anio}-{mes:02d}-Q{quincena}"
    fecha_ini = date(anio, mes, 1) if quincena == 1 else date(anio, mes, 16)
    fecha_fin = date(anio, mes, 15) if quincena == 1 else date(anio, mes, monthrange(anio, mes)[1])
    
    bucket_name = os.getenv('B2_BUCKET_NAME')
    nombre_pdf = f"reportes/{anio}/{mes:02d}/Q{quincena}/{id_imss}_FIRMADO.pdf"
    nombre_excel = f"reportes/{anio}/{mes:02d}/Q{quincena}/{id_imss}_DATOS.xlsx"

    reporte_existente = db.query(models.ReporteQuincenal).filter(
        models.ReporteQuincenal.id_imss == id_imss,
        models.ReporteQuincenal.quincena == periodo_str
    ).first()   

    if reporte_existente:
        if reporte_existente.estado == models.EstadoReporte.APROBADO:
            raise HTTPException(
                status_code=400, 
                detail=f"La quincena {periodo_str} ya fue APROBADA por el Coordinador Estatal. Si necesitas corregirla, contáctalo directamente."
            )
        else:
            # Borramos los archivos viejos de Backblaze PRIMERO
            if reporte_existente.url_documento:
                try: s3_client.delete_object(Bucket=bucket_name, Key=reporte_existente.url_documento)
                except: pass
            if getattr(reporte_existente, 'url_excel', None):
                try: s3_client.delete_object(Bucket=bucket_name, Key=reporte_existente.url_excel)
                except: pass
            
            # Borramos de la BD las asistencias y el reporte viejo
            db.delete(reporte_existente)
            db.query(models.PeasAsistencia).filter(
                models.PeasAsistencia.id_imss == id_imss,
                func.date(models.PeasAsistencia.fecha_hora) >= fecha_ini,
                func.date(models.PeasAsistencia.fecha_hora) <= fecha_fin
            ).delete(synchronize_session=False)
            db.commit()

    # 3. Ahora sí, subimos los archivos NUEVOS a la nube con el camino totalmente libre
    try:
        excel_contents = await archivo_excel.read() # Leemos el excel primero para tenerlo listo
        
        # Subir PDF
        s3_client.upload_fileobj(archivo_pdf.file, bucket_name, nombre_pdf, ExtraArgs={"ContentType": "application/pdf"})
        
        # Subir Excel
        s3_client.upload_fileobj(BytesIO(excel_contents), bucket_name, nombre_excel, ExtraArgs={"ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al subir archivos a la nube: {str(e)}")

    # 4. Crear el nuevo registro principal en la BD
    nuevo_reporte = models.ReporteQuincenal(
        id_imss=id_imss,
        quincena=periodo_str,
        fecha_inicio=fecha_ini,
        fecha_fin=fecha_fin,
        url_documento=nombre_pdf,
        url_excel=nombre_excel,
        subido_por=subido_por
    )
    db.add(nuevo_reporte)

    # 4. Procesar el Excel y crear las Asistencias
    df = pd.read_excel(BytesIO(excel_contents))
    df = df.dropna(how='all')
    mx_tz = pytz.timezone('America/Mexico_City')
    
    asistencias_a_guardar = []
    dias_validos = 0 

    for index, row in df.iterrows():
        if pd.isna(row.get('FECHA (DD/MM/AAAA)')):
            continue

        fecha_cruda = pd.to_datetime(row['FECHA (DD/MM/AAAA)'], dayfirst=True)
        
        # Leemos los valores crudos exactamente igual que en tu previsualizador
        entrada_val = row.get('HORA_ENTRADA (HH:MM)')
        salida_val = row.get('HORA_SALIDA (HH:MM)')

        # Validamos de forma segura con Pandas
        hora_entrada = str(entrada_val).strip() if pd.notnull(entrada_val) and str(entrada_val).strip() != "" else "--:--"
        hora_salida = str(salida_val).strip() if pd.notnull(salida_val) and str(salida_val).strip() != "" else "--:--"

        # Sumamos 1 día si hay registro de entrada o salida
        if hora_entrada != "--:--" or hora_salida != "--:--":
            dias_validos += 1

        # Formateo y guardado de Entradas
        if hora_entrada != "--:--":
            entrada_limpia = hora_entrada[:5] # Extraemos solo HH:MM por seguridad
            dt_entrada_local = mx_tz.localize(datetime.strptime(f"{fecha_cruda.strftime('%Y-%m-%d')} {entrada_limpia}", "%Y-%m-%d %H:%M"))
            dt_entrada_utc = dt_entrada_local.astimezone(pytz.utc).replace(tzinfo=None)
            asistencias_a_guardar.append(models.PeasAsistencia(id_imss=id_imss, tipo="Entrada", fecha_hora=dt_entrada_utc))

        # Formateo y guardado de Salidas
        if hora_salida != "--:--":
            salida_limpia = hora_salida[:5]
            dt_salida_local = mx_tz.localize(datetime.strptime(f"{fecha_cruda.strftime('%Y-%m-%d')} {salida_limpia}", "%Y-%m-%d %H:%M"))
            dt_salida_utc = dt_salida_local.astimezone(pytz.utc).replace(tzinfo=None)
            asistencias_a_guardar.append(models.PeasAsistencia(id_imss=id_imss, tipo="Salida", fecha_hora=dt_salida_utc))

    # Guardamos todas las asistencias de golpe (Bulk Insert)
    if asistencias_a_guardar:
        db.add_all(asistencias_a_guardar)

    # 👇 LE ASIGNAMOS EL CONTEO EXACTO
    nuevo_reporte.total_dias = dias_validos

    # Confirmamos todos los cambios en la base de datos
    db.commit()

    return {
        "mensaje": "Archivos subidos y asistencias registradas exitosamente.",
        "dias_procesados": dias_validos
    }

@router.get("/reporte-quincenal/ver-documento", tags=["Encargado Unidad"])
async def obtener_url_documento(ruta: str):
    """
    Genera una URL temporal y segura (válida por 1 hora) 
    para ver un documento almacenado en Backblaze B2.
    """
    if not ruta:
        raise HTTPException(status_code=400, detail="Ruta del documento no proporcionada")
        
    try:
        bucket_name = os.getenv('B2_BUCKET_NAME')
        
        # Generamos la "Presigned URL" con boto3
        url_temporal = s3_client.generate_presigned_url(
            ClientMethod='get_object',
            Params={
                'Bucket': bucket_name,
                'Key': ruta
            },
            ExpiresIn=3600 # El link expira en 1 hora (3600 segundos)
        )
        
        return {"url": url_temporal}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar enlace seguro: {str(e)}")

@router.get("/coordinador/generar-formato2/{entidad}/{quincena}", tags=["Coordinador Estatal"])
async def obtener_datos_formato_2(entidad: str, quincena: str, db: Session = Depends(get_db)):
    entidad = entidad.upper().strip()
    
    # 1. Buscamos todos los registros validados de ese estado en esa quincena
    registros = db.query(models.BitacoraEstatalValidada).filter(
        models.BitacoraEstatalValidada.entidad == entidad,
        models.BitacoraEstatalValidada.quincena_validada == quincena
    ).order_by(models.BitacoraEstatalValidada.profesional_salud).all()
    
    if not registros:
        raise HTTPException(status_code=404, detail="No hay registros validados para esta quincena.")

    # 2. Calculamos el total de días de participación
    total_dias = sum(reg.dias_participacion for reg in registros if reg.dias_participacion)

    # 3. Formateamos la respuesta para que React la consuma fácil
    detalle = []
    for i, reg in enumerate(registros, start=1):
        detalle.append({
            "no": i,
            "id_imb": reg.id_imss,
            "nombre": reg.profesional_salud,
            "especialidad": reg.especialidad,
            "turno": reg.turno,
            "clues": reg.clues_ib,
            "unidad": reg.unidad_medica,
            "dias": reg.dias_participacion
        })

    return {
        "entidad": entidad,
        "quincena": quincena,
        "total_dias": total_dias,
        "medicos": detalle
    }

@router.post("/coordinador/subir-formato-estatal", tags=["Coordinador Estatal"])
async def subir_formato_estatal(
    entidad: str = Form(...),
    anio: int = Form(...),
    mes: int = Form(...),
    quincena: int = Form(...),
    subido_por: str = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not archivo.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="El formato final debe ser un archivo PDF escaneado.")

    entidad = entidad.upper().strip()
    periodo_str = f"{anio}-{mes:02d}-Q{quincena}"

    # Verificamos si ya existe uno subido para reemplazarlo
    formato_existente = db.query(models.FormatoEstatalFirmado).filter(
        models.FormatoEstatalFirmado.entidad == entidad,
        models.FormatoEstatalFirmado.quincena == periodo_str
    ).first()

    # Generar ruta única en Backblaze
    bucket_name = os.getenv('B2_BUCKET_NAME')
    nombre_unico = f"formatos_estatales/{anio}/{mes:02d}/Q{quincena}/{entidad}_FORMATO2.pdf"

    try:
        # 1. SI YA HABÍA UNO VIEJO, LO BORRAMOS DE B2 PRIMERO 
        # (Esto libera el nombre exacto en la nube antes de subir el nuevo)
        if formato_existente and formato_existente.url_documento:
            try: 
                s3_client.delete_object(Bucket=bucket_name, Key=formato_existente.url_documento)
            except Exception: 
                pass

        # 2. SUBIMOS EL NUEVO ARCHIVO CON EL CAMINO TOTALMENTE LIBRE
        s3_client.upload_fileobj(
            archivo.file, bucket_name, nombre_unico, ExtraArgs={"ContentType": "application/pdf"}
        )
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al subir a la nube: {str(e)}")

    if formato_existente:
        # Actualizamos el existente
        formato_existente.url_documento = nombre_unico
        formato_existente.fecha_subida = func.now()
        formato_existente.subido_por = subido_por
        formato_existente.estado = models.EstadoReporte.PENDIENTE 
        formato_existente.observaciones = None
    else:
        # Creamos uno nuevo
        nuevo_formato = models.FormatoEstatalFirmado(
            entidad=entidad,
            quincena=periodo_str,
            url_documento=nombre_unico,
            subido_por=subido_por
        )
        db.add(nuevo_formato)

    db.commit()

    return {"mensaje": "Formato Estatal Firmado guardado con éxito", "url": nombre_unico}

@router.get("/coordinador/historial-formatos/{entidad}", tags=["Coordinador Estatal"])
async def obtener_historial_formatos(entidad: str, db: Session = Depends(get_db)):
    entidad_limpia = entidad.upper().strip()
    
    # Traemos los formatos ordenados del más reciente al más antiguo
    formatos = db.query(models.FormatoEstatalFirmado).filter(
        models.FormatoEstatalFirmado.entidad == entidad_limpia
    ).order_by(desc(models.FormatoEstatalFirmado.fecha_subida)).all()
    
    historial = []
    for f in formatos:
        historial.append({
            "id": f.id,
            "quincena": f.quincena,
            "fecha_subida": formatear_fecha_local(f.fecha_subida),
            "subido_por": f.subido_por,
            "url_documento": f.url_documento,
            "estado": f.estado.value if hasattr(f.estado, 'value') else f.estado, 
            "observaciones": f.observaciones 
        })
        
    return historial

@router.get("/nacional/formato3/{anio}/{mes}/{quincena}", tags=["Nacional"])
async def obtener_formato_3_nacional(anio: int, mes: int, quincena: str, db: Session = Depends(get_db)):
    
    # 1. Configuración dinámica del texto para el Excel
    meses_nombres = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    nombre_mes = meses_nombres[mes]
    ultimo_dia = calendar.monthrange(anio, mes)[1] # Detecta automáticamente si el mes trae 28, 30 o 31 días
    
    if quincena == "completo":
        periodos = [f"{anio}-{mes:02d}-Q1", f"{anio}-{mes:02d}-Q2"]
        texto_periodo_excel = f"1 AL {ultimo_dia} DE {nombre_mes} DE {anio}"
    elif quincena == "1":
        periodos = [f"{anio}-{mes:02d}-Q1"]
        texto_periodo_excel = f"1 AL 15 DE {nombre_mes} DE {anio}"
    elif quincena == "2":
        periodos = [f"{anio}-{mes:02d}-Q2"]
        texto_periodo_excel = f"16 AL {ultimo_dia} DE {nombre_mes} DE {anio}"
    else:
        raise HTTPException(status_code=400, detail="Quincena inválida")
    
    # 2. Buscamos formatos APROBADOS
    formatos_aprobados = db.query(models.FormatoEstatalFirmado).filter(
        models.FormatoEstatalFirmado.quincena.in_(periodos),
        models.FormatoEstatalFirmado.estado == models.EstadoReporte.APROBADO
    ).all()
    
    if not formatos_aprobados:
        raise HTTPException(status_code=400, detail="No se encontraron registros aprobados para este periodo.")

    # 3. Candado de seguridad
    condiciones_aprobadas = [
        and_(
            models.BitacoraEstatalValidada.entidad == f.entidad,
            models.BitacoraEstatalValidada.quincena_validada == f.quincena
        ) for f in formatos_aprobados
    ]

    # 4. LA MAGIA ESTÁ AQUÍ: Agrupamos por médico y sumamos los días
    registros = db.query(
        models.BitacoraEstatalValidada.entidad,
        models.BitacoraEstatalValidada.unidad_medica,
        models.BitacoraEstatalValidada.clues_ib,
        models.BitacoraEstatalValidada.especialidad,
        models.BitacoraEstatalValidada.turno,
        models.BitacoraEstatalValidada.profesional_salud,
        func.sum(models.BitacoraEstatalValidada.dias_participacion).label("dias_totales") # Sumamos los días
    ).filter(
        or_(*condiciones_aprobadas)
    ).group_by(
        # Le decimos a SQL que agrupe todas las filas que compartan estos datos idénticos
        models.BitacoraEstatalValidada.entidad,
        models.BitacoraEstatalValidada.unidad_medica,
        models.BitacoraEstatalValidada.clues_ib,
        models.BitacoraEstatalValidada.especialidad,
        models.BitacoraEstatalValidada.turno,
        models.BitacoraEstatalValidada.profesional_salud
    ).order_by(
        models.BitacoraEstatalValidada.entidad, 
        models.BitacoraEstatalValidada.unidad_medica
    ).all()
    
    # 5. Formateamos la respuesta
    detalle = []
    for i, reg in enumerate(registros, start=1):
        detalle.append({
            "no": i,
            "unidad": reg.unidad_medica or "SIN UNIDAD",
            "entidad": reg.entidad,
            "clues": reg.clues_ib,
            "especialidad": reg.especialidad,
            "turno": reg.turno,
            "medico": reg.profesional_salud,
            "dias": reg.dias_totales # Inyectamos la suma correcta
        })
        
    return {
        "periodo_texto": texto_periodo_excel, # Pasamos el string perfecto para tu título
        "anio": anio,
        "mes": mes,
        "quincena": quincena,
        "medicos": detalle
    }

@router.get("/nacional/formato4/{anio}/{mes}/{quincena}", tags=["Nacional"])
async def obtener_formato_4_nacional(anio: int, mes: int, quincena: str, db: Session = Depends(get_db)):
    
    # 1. Configuración dinámica del texto y periodos
    meses_nombres = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    nombre_mes = meses_nombres[mes]
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    
    if quincena == "completo":
        periodos = [f"{anio}-{mes:02d}-Q1", f"{anio}-{mes:02d}-Q2"]
        texto_periodo_excel = f"1 AL {ultimo_dia} DE {nombre_mes} DE {anio}"
    elif quincena == "1":
        periodos = [f"{anio}-{mes:02d}-Q1"]
        texto_periodo_excel = f"1 AL 15 DE {nombre_mes} DE {anio}"
    elif quincena == "2":
        periodos = [f"{anio}-{mes:02d}-Q2"]
        texto_periodo_excel = f"16 AL {ultimo_dia} DE {nombre_mes} DE {anio}"
    else:
        raise HTTPException(status_code=400, detail="Quincena inválida")
    
    # 2. Buscamos qué formatos están APROBADOS
    formatos_aprobados = db.query(models.FormatoEstatalFirmado).filter(
        models.FormatoEstatalFirmado.quincena.in_(periodos),
        models.FormatoEstatalFirmado.estado == models.EstadoReporte.APROBADO
    ).all()
    
    if not formatos_aprobados:
        raise HTTPException(
            status_code=400, 
            detail="No se puede generar el resumen: Aún no hay registros APROBADOS para este periodo."
        )
    
    # 3. Armamos el mismo candado de seguridad que en el formato 3
    condiciones_aprobadas = [
        and_(
            models.BitacoraEstatalValidada.entidad == f.entidad,
            models.BitacoraEstatalValidada.quincena_validada == f.quincena
        ) for f in formatos_aprobados
    ]
    
    # 4. Le pedimos a PostgreSQL que cuente los médicos (SIN REPETIR) y sume los días
    resultados = db.query(
        models.BitacoraEstatalValidada.entidad,
        func.count(distinct(models.BitacoraEstatalValidada.profesional_salud)).label("total_medicos"), # <-- CRÍTICO: Cuenta médicos únicos
        func.sum(models.BitacoraEstatalValidada.dias_participacion).label("total_dias")
    ).filter(
        or_(*condiciones_aprobadas)
    ).group_by(
        models.BitacoraEstatalValidada.entidad
    ).order_by(
        models.BitacoraEstatalValidada.entidad
    ).all()
    
    detalle = []
    gran_total_medicos = 0
    gran_total_dias = 0
    
    for res in resultados:
        dias_validos = res.total_dias or 0
        detalle.append({
            "entidad": res.entidad,
            "medicos": res.total_medicos,
            "dias": dias_validos
        })
        gran_total_medicos += res.total_medicos
        gran_total_dias += dias_validos
        
    return {
        "periodo_texto": texto_periodo_excel, # Mismo título exacto que el formato 3
        "anio": anio,
        "mes": mes,
        "quincena": quincena,
        "resumen": detalle,
        "gran_total_medicos": gran_total_medicos,
        "gran_total_dias": gran_total_dias
    }

@router.post("/nacional/subir-formato-nacional", tags=["Nacional"])
async def subir_formato_nacional(
    anio: int = Form(...),
    mes: int = Form(...),
    quincena: str = Form(...), # <-- Cambiado a str
    subido_por: str = Form(...),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not archivo.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="El formato final debe ser un archivo PDF.")

    # Normalizamos el periodo para que coincida con la lógica que ya usamos
    periodo_str = f"{anio}-{mes:02d}-Q{quincena}"
    
    bucket_name = os.getenv('B2_BUCKET_NAME')
    # Ajustamos la ruta para que soporte el string "completo" en el nombre del archivo
    nombre_unico = f"formatos_nacionales/{anio}/{mes:02d}/{quincena}/NACIONAL_FORMATO3y4.pdf"

    # 1. Verificamos si existe en la BD
    formato_existente = db.query(models.FormatoNacionalFirmado).filter(
        models.FormatoNacionalFirmado.quincena == periodo_str
    ).first()

    # 2. Borrado seguro en la nube
    if formato_existente and formato_existente.url_documento:
        try:
            s3_client.delete_object(Bucket=bucket_name, Key=formato_existente.url_documento)
        except Exception:
            pass

    # 3. Subida a la nube
    try:
        s3_client.upload_fileobj(
            archivo.file, 
            bucket_name, 
            nombre_unico, 
            ExtraArgs={"ContentType": "application/pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al subir a la nube: {str(e)}")

    # 4. Actualización de BD
    if formato_existente:
        formato_existente.url_documento = nombre_unico
        formato_existente.fecha_subida = func.now()
        formato_existente.subido_por = subido_por
    else:
        nuevo_formato = models.FormatoNacionalFirmado(
            quincena=periodo_str, 
            url_documento=nombre_unico, 
            subido_por=subido_por
        )
        db.add(nuevo_formato)

    db.commit()
    return {"mensaje": "Formato Nacional Firmado guardado con éxito"}

@router.get("/nacional/historial-formatos", tags=["Nacional"])
async def obtener_historial_nacional(db: Session = Depends(get_db)):
    # Traemos todos los formatos nacionales ordenados por fecha
    formatos = db.query(models.FormatoNacionalFirmado).order_by(
        desc(models.FormatoNacionalFirmado.fecha_subida)
    ).all()
    
    historial = []
    for f in formatos:
        historial.append({
            "id": f.id,
            "quincena": f.quincena,
            "fecha_subida": formatear_fecha_local(f.fecha_subida),
            "subido_por": f.subido_por,
            "url_documento": f.url_documento
        })
        
    return historial

@router.get("/reporte-quincenal/estado-subidos/{anio}/{mes}/{quincena}", tags=["Responsable Unidad"])
async def obtener_estado_subidos(anio: str, mes: str, quincena: str, db: Session = Depends(get_db)):
    mes_formateado = mes.zfill(2) 
    periodo_buscado = f"{anio}-{mes_formateado}-Q{quincena}"
    
    # 1. Hacemos un JOIN con models.Doctor para traer sus datos
    reportes = db.query(models.ReporteQuincenal, models.Doctor)\
        .outerjoin(models.Doctor, models.ReporteQuincenal.id_imss == models.Doctor.id_imss)\
        .filter(models.ReporteQuincenal.quincena == periodo_buscado)\
        .all()
    
    estado_reportes = {}
    for rep, doc in reportes:
        # 2. Armamos el nombre completo aquí en el servidor
        if doc:
            nombre_completo = f"{doc.nombre} {doc.apellido_paterno or ''} {doc.apellido_materno or ''}".strip()
        else:
            nombre_completo = f"Médico ID: {rep.id_imss}"

        # 3. Extraemos el valor del enum e inyectamos el nombre
        estado_reportes[rep.id_imss] = {
            "estado": rep.estado.value if hasattr(rep.estado, 'value') else rep.estado,
            "observaciones": rep.observaciones,
            "nombre_medico": nombre_completo # <-- AQUÍ VA EL NOMBRE REAL
        }
        
    return estado_reportes

@router.post("/reporte-quincenal/rechazar/{id_reporte}", tags=["Coordinador Estatal"])
async def rechazar_reporte_unidad(id_reporte: int, req: RechazoRequest, db: Session = Depends(get_db)):
    # 1. Buscamos el reporte por su ID
    reporte = db.query(models.ReporteQuincenal).filter(models.ReporteQuincenal.id == id_reporte).first()
    if not reporte:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")
    
    # 2. Le cambiamos el semáforo y le guardamos el texto
    reporte.estado = models.EstadoReporte.RECHAZADO
    reporte.observaciones = req.observaciones
    
    db.commit()
    return {"mensaje": "Reporte rechazado. El Encargado de Unidad ha sido notificado."}

@router.post("/coordinador/rechazar-formato-estatal/{id_formato}", tags=["Nacional"])
async def rechazar_formato_estatal(id_formato: int, req: RechazoRequest, db: Session = Depends(get_db)):
    # 1. Buscamos el formato estatal por su ID
    formato = db.query(models.FormatoEstatalFirmado).filter(models.FormatoEstatalFirmado.id == id_formato).first()
    if not formato:
        raise HTTPException(status_code=404, detail="Formato estatal no encontrado")
    
    # 2. Le cambiamos el semáforo y le guardamos el texto
    formato.estado = models.EstadoReporte.RECHAZADO
    formato.observaciones = req.observaciones
    
    db.commit()
    return {"mensaje": "Formato estatal rechazado. El Coordinador ha sido notificado."}

@router.get("/nacional/estado-formatos/{anio}/{mes}/{quincena}", tags=["Nacional"])
async def obtener_estado_formatos_estatales(anio: int, mes: str, quincena: str, db: Session = Depends(get_db)):
    # Nota: cambiamos quincena a str en los parámetros
    mes_formateado = str(mes).zfill(2)
    
    # 1. Determinamos los periodos a buscar
    if quincena == "completo":
        periodos = [f"{anio}-{mes_formateado}-Q1", f"{anio}-{mes_formateado}-Q2"]
    elif quincena in ["1", "2"]:
        periodos = [f"{anio}-{mes_formateado}-Q{quincena}"]
    else:
        raise HTTPException(status_code=400, detail="Quincena inválida")
    
    # 2. Traemos todos los Formatos subidos en esos periodos (puede ser uno o ambos)
    formatos = db.query(models.FormatoEstatalFirmado).filter(
        models.FormatoEstatalFirmado.quincena.in_(periodos)
    ).order_by(
        models.FormatoEstatalFirmado.entidad,
        models.FormatoEstatalFirmado.quincena
    ).all()
    
    # 3. Devolvemos la lista limpia
    resultado = []
    for f in formatos:
        # Truco visual: Si están viendo el mes completo, le agregamos (Q1) o (Q2) al nombre
        # para que no se confundan viendo estados duplicados en la tabla.
        sufijo_quincena = f.quincena[-2:] # Extrae "Q1" o "Q2" del string "2026-08-Q1"
        nombre_entidad = f"{f.entidad} ({sufijo_quincena})" if quincena == "completo" else f.entidad

        resultado.append({
            "id": f.id,
            "entidad": nombre_entidad,  # <-- Aquí inyectamos el nombre ajustado
            "url_documento": f.url_documento,
            "estado": f.estado.value if hasattr(f.estado, 'value') else f.estado,
            "observaciones": f.observaciones,
            "fecha_subida": formatear_fecha_local(f.fecha_subida)
        })
        
    return resultado

@router.post("/nacional/aprobar-formato/{id_formato}", tags=["Nacional"])
async def aprobar_formato_estatal(id_formato: int, db: Session = Depends(get_db)):
    formato = db.query(models.FormatoEstatalFirmado).filter(models.FormatoEstatalFirmado.id == id_formato).first()
    if not formato:
        raise HTTPException(status_code=404, detail="Formato estatal no encontrado")
    
    formato.estado = models.EstadoReporte.APROBADO
    formato.observaciones = None # Limpiamos cualquier observación vieja
    
    db.commit()
    return {"mensaje": "Formato estatal aprobado y bloqueado para su generación."}

@router.get("/coordinador/reportes-validados/{entidad}/{periodo}", tags=["Coordinador Estatal"])
async def obtener_reportes_validados(entidad: str, periodo: str, db: Session = Depends(get_db)):
    entidad_limpia = entidad.upper().strip()
    
    # Traemos los registros de la bitácora estatal validada para ese periodo y entidad
    registros = db.query(models.BitacoraEstatalValidada).filter(
        models.BitacoraEstatalValidada.entidad == entidad_limpia,
        models.BitacoraEstatalValidada.quincena_validada == periodo
    ).order_by(models.BitacoraEstatalValidada.profesional_salud).all()
    
    validados = []
    for reg in registros:
        reporte_orig = db.query(models.ReporteQuincenal).filter(
            models.ReporteQuincenal.id == reg.id_reporte_quincenal
        ).first()
        validados.append({
            "id_bitacora": reg.id,
            "id_reporte": reg.id_reporte_quincenal,
            "quincena": reg.quincena_validada,
            "clues": reg.clues_ib,
            "unidad": reg.unidad_medica,
            "medico": reg.profesional_salud,
            "id_imss": reg.id_imss,
            "especialidad": reg.especialidad,
            "turno": reg.turno,
            "dias": reg.dias_participacion,
            "url_pdf": reporte_orig.url_documento if reporte_orig else None
        })
        
    return validados

class RevocarRequest(BaseModel):
    observaciones: str

@router.post("/coordinador/revocar-reporte/{id_reporte}", tags=["Coordinador Estatal"])
async def revocar_reporte(id_reporte: int, req: RevocarRequest, db: Session = Depends(get_db)):
    # 1. Buscamos el reporte original que la unidad subió
    reporte_original = db.query(models.ReporteQuincenal).filter(models.ReporteQuincenal.id == id_reporte).first()
    if not reporte_original:
        raise HTTPException(status_code=404, detail="Reporte original no encontrado")
        
    # 2. Eliminamos el registro validado de la bitácora estatal
    db.query(models.BitacoraEstatalValidada).filter(
       models.BitacoraEstatalValidada.id_reporte_quincenal == id_reporte
    ).delete()
    
    # 3. Regresamos el reporte original a estado RECHAZADO y guardamos el regaño
    reporte_original.estado = models.EstadoReporte.RECHAZADO
    reporte_original.observaciones = req.observaciones
    
    db.commit()
    return {"mensaje": "Reporte revocado exitosamente y devuelto a la unidad."}

@router.get("/reporte-quincenal/descargar-formato-impresion", tags=["Reportes PEAS"])
async def descargar_formato_impresion(anio: int, mes: int, quincena: int):
    # 1. Cargamos tu plantilla base
    directorio_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_plantilla = os.path.join(directorio_actual, "plantilla_firma_manual.xlsx")
    
    if not os.path.exists(ruta_plantilla):
        raise HTTPException(status_code=404, detail="Plantilla base no encontrada.")

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active

    # 2. Configuramos el rango de días según el mes y quincena elegidos
    meses = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    nombre_mes = meses[mes]

    if quincena == 1:
        dia_inicio = 1
        dia_fin = 15
    else:
        dia_inicio = 16
        dia_fin = calendar.monthrange(anio, mes)[1] 

    # 3. Escribir el "Periodo" dinámico en el encabezado
    # SUSTITUYE 'G4' por la celda real donde está tu texto de "Periodo: 16 al 30..."
    ws['I4'] = f"{dia_inicio} al {dia_fin} de {nombre_mes} de {anio}"

    # 4. Escribir las fechas fila por fila
    fila_inicio = 8 
    dias_totales = range(dia_inicio, dia_fin + 1)
    
    for i in range(16): # 16 es el máximo de filas necesarias (ej. meses con 31 días en Q2)
        celda = ws.cell(row=fila_inicio + i, column=2) # Asumiendo que la Fecha está en la Columna A (1)
        
        if i < len(dias_totales):
            dia = dias_totales[i]
            # Formato estricto DD/MM/AAAA
            celda.value = f"{dia:02d}/{mes:02d}/{anio}"
        else:
            # Si sobran filas (ej. la primera quincena solo usa 15), las dejamos en blanco
            celda.value = ""

    # 5. Guardar en memoria virtual (StreamingResponse no gasta disco de tu servidor)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"Registro_Firmas_{nombre_mes}_Q{quincena}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )

@router.get("/reporte-quincenal/descargar-plantilla-dinamica", tags=["Reportes PEAS"])
async def descargar_plantilla_dinamica(anio: int, mes: int, quincena: int):
    directorio_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_plantilla = os.path.join(directorio_actual, "formato_asistencias.xlsx")
    
    if not os.path.exists(ruta_plantilla):
        raise HTTPException(status_code=404, detail="Plantilla base no encontrada.")

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active

    if quincena == 1:
        dia_inicio = 1
        dia_fin = 15
    else:
        dia_inicio = 16
        dia_fin = calendar.monthrange(anio, mes)[1]

    dias_totales = range(dia_inicio, dia_fin + 1)
    fila_inicio = 2 
    
    for i in range(16): 
        celda = ws.cell(row=fila_inicio + i, column=1) 
        
        if i < len(dias_totales):
            dia = dias_totales[i]
            celda.value = f"{dia:02d}/{mes:02d}/{anio}"
        else:
            celda.value = ""

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"Plantilla_Asistencia_{anio}_{mes:02d}_Q{quincena}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )

